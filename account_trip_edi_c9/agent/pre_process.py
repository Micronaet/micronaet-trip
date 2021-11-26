# -*- coding: utf-8 -*-
###############################################################################
#
#    Copyright (C) 2001-2014 Micronaet SRL (<http://www.micronaet.it>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as published
#    by the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################
import pdb
import sys
import os
import shutil
import openpyxl
import ConfigParser
from datetime import datetime


# -----------------------------------------------------------------------------
# Utility:
# -----------------------------------------------------------------------------
def sort_line(row):
    """ Generate order depend on start code, the row is the string generated
        for ERP (code 4th field)
    """
    code = row.split('|')[11].upper()
    start_1 = code[:1]
    start_2 = code[:2]
    start_3 = code[:3]

    # -------------------------------------------------------------------------
    # 3 char test start:
    # -------------------------------------------------------------------------
    if start_3 == 'SPA':
        return 4, code  # pasta

    # -------------------------------------------------------------------------
    # 2 char test start:
    # -------------------------------------------------------------------------
    if start_2 in ('SP', 'SS'):
        return 1, code  # freeze

    # -------------------------------------------------------------------------
    # 1 char test start:
    # -------------------------------------------------------------------------
    if start_1 in 'CDFPSV':
        return 1, code  # freeze
    elif start_1 in 'BILT':
        return 2, code  # fresh
    elif start_1 in 'HO':
        return 3, code  # dry
    elif start_1 in 'G':
        return 4, code  # pasta
    else: # Error list
        f_error = open('./sort_error.log', 'w')
        f_error.write(
            'Char not found %s - %s -%s\n' % (start_1, start_2, start_3))
        print('Char not found %s - %s - %s' % (start_1, start_2, start_3))
        f_error.close()
        return 5, code


def clean_text(text, length, uppercase=False, error=None, truncate=False):
    """ Return clean text with limit cut
        Log in error if over length
    """
    if error is None:
        error = []

    text = text.strip()
    if len(text) > length:
        if truncate:
            text = text[:length]
        else:
            error.append('Text: %s > %s' % (text, length))
    if uppercase:
        return text.upper()
    return text


def clean_date(company_date):  # todo!!!
    """ Current format: YYMMDD
    """
    return company_date.strip()


def clean_float(
        value, length, decimal=3, multiple=1.0, separator='.', error=None):
    """ Clean float and return float format
    """
    value = value.strip()
    if not value:
        return 0.0

    if error is None:
        error = []
    try:
        value = value.replace(',', '.')
        float_value = float(value.strip())
    except:
        error.append('Not a float: %s' % value)
        float_value = 0.0
    float_value /= multiple
    mask = '%%%s.%sf' % (length, decimal)
    res = mask % float_value
    res = res.replace('.', separator)
    return res


def log_on_file(message, mode='INFO', file_list=None, verbose=True):
    """ Write message in file list passed
        mode: INFO WARNING ERROR
    """
    if file_list is None:
        return False

    message_log = '%s. [%s] %s\n' % (datetime.now(), mode, message)
    if verbose:
        print(message_log.strip())

    for f_log in file_list:
        if not f_log:
            continue
        f_log.write(message_log)
    return True

# -----------------------------------------------------------------------------
# Read configuration parameter:
# -----------------------------------------------------------------------------
try:
    company = sys.argv[1]
    if len(company) > 3:
        print('Run: python ./pre_process.py CMP (must be max 3 char length)')
        sys.exit()
except:
    print('Launch program as:\n\npython ./pre_process.py CMP')
    sys.exit()

try:
    cfg_file = os.path.expanduser('./%s.cfg' % company)
except:
    print('Config file not found: %s.cfg' % company)
    sys.exit()

config = ConfigParser.ConfigParser()
config.read([cfg_file])

# In parameters:
in_check = os.path.expanduser(config.get('in', 'check'))
in_path = os.path.expanduser(config.get('in', 'path'))
in_history = os.path.expanduser(config.get('in', 'history'))
in_log = os.path.expanduser(config.get('in', 'log'))
in_schedule = os.path.expanduser(config.get('in', 'schedule'))

# Out parameters:
out_check = os.path.expanduser(config.get('out', 'check'))
out_path = os.path.expanduser(config.get('out', 'path'))
out_log = os.path.expanduser(config.get('out', 'log'))
out_schedule = os.path.expanduser(config.get('out', 'schedule'))
out_original = os.path.expanduser(config.get('out', 'original'))

# Mail parameters:
mail_error = config.get('mail', 'error')
mail_info = config.get('mail', 'info')

# Calculated parameters:
f_in_schedule = open(in_schedule, 'a')
f_in_log = open(in_log, 'a')

start = 2
extension = 'xlsx'

f_out_schedule = open(out_schedule, 'a')
f_out_log = open(out_log, 'a')

# -----------------------------------------------------------------------------
# Check mount folder:
# -----------------------------------------------------------------------------
if in_check:
    if not os.path.exists(in_check):
        log_on_file(
            'Cannot mount IN folder: %s' % in_path, mode='ERROR', file_list=[
                f_in_schedule, f_out_schedule])
        sys.exit()

if out_check:
    if not os.path.exists(out_check):
        log_on_file(
            'Cannot mount OUT folder: %s' % out_path, mode='ERROR', file_list=[
                f_in_schedule, f_out_schedule])
        sys.exit()

# -----------------------------------------------------------------------------
# Read IN folder:
# -----------------------------------------------------------------------------
log_on_file(
    'Start import order mode: %s' % company, mode='INFO', file_list=[
        f_in_schedule, f_out_schedule])
for root, dirs, files in os.walk(in_path):
    log_on_file(
        'Read root folder: %s [%s]' % (root, company),
        mode='INFO',
        file_list=[f_in_log, f_out_log])

    for f in files:
        if not f.lower().endswith(extension):
            log_on_file(
                'Jump file not managed: %s' % f, mode='WARNING',
                file_list=[
                    f_in_schedule, f_out_schedule])
            continue
        command = 'NEW'  # Only this!

        # Fullname needed:
        file_in = os.path.join(root, f)
        file_history = os.path.join(in_history, f)
        file_original = os.path.join(out_original, f)

        create_date = datetime.fromtimestamp(
            os.path.getctime(file_in)
            ).strftime('%Y%m%d_%H%M%S')

        file_out = os.path.join(
            out_path, '%s' % (
                '%s_%s_%s.txt' % (
                    create_date, f[:-4], command),
                ))

        # ---------------------------------------------------------------------
        #                    Read Excel files:
        # ---------------------------------------------------------------------
        error = []
        try:
           wb = openpyxl.load_workbook(file_in)
        except:
            print('Errore: %s' % (sys.exc_info(), ))
            log_on_file(
                'Error opening file: %s [%s]' % (file_in, company),
                mode='ERROR', file_list=[f_in_log, f_out_log])
            sys.exit()

        ws = wb.active
        row_out = []
        log_on_file(
            'Parsing file: %s [%s]' % (file_in, company),
            mode='INFO', file_list=[f_in_log, f_out_log])

        pdb.set_trace()
        for row in range(start, ws.max_row):
            record = {
                # Header information:
                'order': ws.cell(row, 1).value,
                'serie': ws.cell(row, 2).value,
                'date': ws.cell(row, 3).value,
                # Ragione sociale 4

                # Stock inforation:
                'stock_code': ws.cell(row, 5).value,
                'stock_name': ws.cell(row, 6).value,
                'stock_address': ws.cell(row, 7).value,
                'stock_cap': ws.cell(row, 8).value,
                'stock_city': ws.cell(row, 9).value,
                'stock_province': ws.cell(row, 10).value,

                # Detail information:
                'customer_code': ws.cell(row, 11).value,
                'customer_name': ws.cell(row, 12).value,
                'customer_extra': ws.cell(row, 13).value,

                'company_code': ws.cell(row, 14).value,
                'uom': ws.cell(row, 15).value,
                'quantity': ws.cell(row, 16).value,
                'price': ws.cell(row, 17).value,
                'subtotal': ws.cell(row, 18).value,
                'deadline': ws.cell(row, 19).value,
                'line_note': ws.cell(row, 20).value,
                'delivery_note': ws.cell(row, 21).value,
                'cig': ws.cell(row, 22).value,
                'cup': ws.cell(row, 23).value,
                'calendar': ws.cell(row, 24).value,
            }

            # -----------------------------------------------------------------
            # Convert row input file:
            # -----------------------------------------------------------------
            row_out.append(
                'MRK|%-10s|%-10s|%-10s|%-8s|'
                '%-13s|%4s|%-16s|%-60s|%-2s|%15s|'
                '%-16s|%-60s|%-2s|%15s|'
                '%-8s|%-40s|%-40s|%-15s\r\n' % (
                    # Destination code:
                    '',
                    clean_text(
                        record['stock_code'], 10, error=error,
                        truncate=True),
                    '',

                    clean_date(record['deadline']),
                    clean_text(
                        record['order'], 13, error=error, truncate=True),

                    # Line:
                    row,  # remove header line
                    clean_text(record['company_code'], 16, error=error,
                               truncate=True, uppercase=True),
                    clean_text(record['customer_name'], 60, error=error,
                               truncate=True),
                    clean_text(record['uom'], 2, error=error, truncate=True,
                               uppercase=True),
                    clean_float(
                        record['quantity'], 15, 2, error=error),

                    # Supplier reference (here is the same)
                    clean_text(record['customer_code'], 16, error=error,
                               truncate=True, uppercase=True),
                    clean_text(record['customer_name'], 60, error=error,
                               truncate=True),
                    clean_text(record['uom'], 2, error=error, truncate=True,
                               uppercase=True),
                    clean_float(
                        record['quantity'], 15, 2, error=error),

                    # Footer:
                    clean_date(record['date']),
                    '',  # header note
                    clean_text(record['line_note'], 40, error=error,
                               truncate=True, uppercase=True),  # line note
                    clean_text(record['cig'], 15, error=error,
                               truncate=True, uppercase=True),  # line note
                    # todo price?
                    ))

        if error:
            log_on_file(
                'Error reading file: %s [%s]' % (error, company),
                mode='ERROR', file_list=[f_in_log, f_out_log])
            continue

        # ---------------------------------------------------------------------
        # Save file out:
        # ---------------------------------------------------------------------
        error = []
        # Convert file:
        f_out = open(file_out, 'w')

        # for line in sorted(row_out, key=lambda x: sort_line(x)):  todo sort?
        for line in row_out:
            f_out.write(line)
        f_out.close()

        if error:
            log_on_file(
                'Error generating out file: %s [%s]' % (error, company),
                mode='ERROR', file_list=[f_in_log, f_out_log])
            continue
        else:
            log_on_file(
                'Created file: %s [%s]' % (file_out, company),
                mode='INFO', file_list=[f_in_log, f_out_log])

        # ---------------------------------------------------------------------
        # History the in file:
        # ---------------------------------------------------------------------
        error = ''
        try:
            os.remove(file_original)
        except:
            pass
        try:
            shutil.copy(file_in, file_original)
            log_on_file(
                'History company file: %s [%s]' % (file_original, company),
                mode='INFO', file_list=[f_in_log, f_out_log])
        except:
            error += 'Error copy original folder %s >> %s' % (
                file_in, file_original)
        try:
            os.remove(file_history)
        except:
            pass
        try:
            shutil.move(file_in, file_history)
            log_on_file(
                'History supplier file: %s [%s]' % (file_history, company),
                mode='INFO', file_list=[f_in_log, f_out_log])
        except:
            error += 'Error moving file %s >> %s' % (file_in, file_history)

        if error:
            log_on_file(
                error, mode='ERROR', file_list=[f_in_log, f_out_log])
            # XXX TODO delete OUT file
            continue
    break  # only root folder is read

log_on_file(
    'End import order mode: %s' % company, mode='INFO', file_list=[
        f_in_schedule, f_out_schedule])

# -----------------------------------------------------------------------------
# History the in file:
# -----------------------------------------------------------------------------
try:
    if f_out_log:
        f_out_log.close()
    if f_in_schedule:
        f_in_schedule.close()
    if f_in_log:
        f_in_log.close()
    if f_out_schedule:
        f_out_schedule.close()
except:
    print('Error closing log file')
